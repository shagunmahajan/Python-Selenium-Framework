import os
import time
import unittest
from pathlib import Path, PurePosixPath
from selenium.webdriver.chrome.options import Options
from selenium.webdriver import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from selenium import webdriver
import random
import openpyxl


# screenshot main directory for problem 2
screenshot_dir = PurePosixPath(Path.cwd()).joinpath("../images/problem2/")

# filename as per execution time
fileName = time.strftime("%Y%m%d-%H%M%S") + ".png"


# driver options
def get_driver():
    chrome_options = Options()
    # chrome_options.add_argument('--headless')
    chrome_options.add_argument('--start-maximized')

    # chrome driver options
    chrome_location = PurePosixPath(Path.cwd()).joinpath("../chromedriver.exe")
    driver = webdriver.Chrome(options=chrome_options, executable_path=chrome_location)

    # firefox driver options
    # firefox_location = PurePosixPath(Path.cwd()).joinpath("../geckodriver.exe")
    # driver = webdriver.Firefox(options=chrome_options, executable_path=firefox_location)

    return driver


# url for the problem, driver selected
url = "http://automationpractice.com/index.php"
driver = get_driver()

action = ActionChains(driver)
wait = WebDriverWait(driver, 10)
budget_amount = 0


# function to take screenshot of the screen
def take_screenshot(screenshot_directory):
    try:
        driver.save_screenshot(screenshot_directory)
        print("Screenshot saved successfully")
    except Exception as e:
        print("Couldn't save the screenshot ")


# load workbook and retrieve data
def excel_data():
    # Loading workbook
    book = openpyxl.load_workbook("../testData.xlsx")
    sheet = book.active
    return sheet


# modify excel data as pr user input
def modify_excel_data(user_data):
    book = openpyxl.load_workbook("../testData.xlsx")
    sheet = book.active
    budget_column_name = sheet.cell(1, 10)
    budget_cell = sheet.cell(2, 10)
    budget_column_name.value = user_data[0]
    budget_cell.value = user_data[1]
    book.save("../testData.xlsx")
    print("Updated excel")


# take user input
def take_user_input():
    budget_column = input("\n Enter column name: ")
    cell_value = input("\nEnter budget value: ")
    return budget_column, cell_value


def tearDown():
    driver.quit()


# function to create account as per excel
def create_account(sheet):
    # add user details
    driver.find_element_by_id('customer_firstname').send_keys(sheet.cell(2, 1).value)
    driver.find_element_by_id('customer_lastname').send_keys(sheet.cell(2, 2).value)
    # driver.find_element_by_id('email').send_keys(sheet.cell(2, 3).value + str(random.randint(1, 1001))+ "@gmail.com")
    driver.find_element_by_id('passwd').send_keys(sheet.cell(2, 4).value)

    # add address details
    driver.find_element_by_id('firstname').send_keys(sheet.cell(2, 1).value)
    driver.find_element_by_id('lastname').send_keys(sheet.cell(2, 2).value)
    driver.find_element_by_id('address1').send_keys(sheet.cell(2, 5).value)
    driver.find_element_by_id('city').send_keys(sheet.cell(2, 6).value)
    driver.find_element_by_id('id_state').send_keys(sheet.cell(2, 7).value)
    driver.find_element_by_id('postcode').send_keys(sheet.cell(2, 8).value)
    driver.find_element_by_id('phone_mobile').send_keys(sheet.cell(2, 9).value)
    driver.find_element_by_id('submitAccount').click()


# unit test case
class SuccessTest(unittest.TestCase):

    # take user input from console and update required column and it's value
    modify_excel_data(take_user_input())

    # open browser
    driver.get(url)

    # click on login button
    driver.find_element_by_class_name("login").click()
    time.sleep(5)

    # retrieve email id and provide input to the email id field using random number generation
    driver.find_element_by_id("email_create").send_keys(
        excel_data().cell(2, 3).value + str(random.randint(1, 1001)) + "@gmail.com")

    # click on create an account button
    driver.find_element_by_id("SubmitCreate").click()

    # adding wait time for page load after clicking on create account button
    time.sleep(8)

    # create account as per excel data
    create_account(excel_data())

    # verify successful creation of account
    navigation_screen = driver.find_element_by_class_name("navigation_page")
    try:
        assert 'My account' in navigation_screen.text
    except Exception as e:
        print("Navigation to My account screen failed")

    driver.find_element_by_class_name('home').click()

    # get popular products
    popular_items_tab = driver.find_element_by_id("homefeatured")

    # scroll to popular products
    driver.execute_script("arguments[0].scrollIntoView();", popular_items_tab)
    popular_items = popular_items_tab.find_elements_by_tag_name('li')
    count = 1

    # add popular products to the cart
    for item in popular_items:
        try:
            # number of items to be added to the cart
            if count > 3:
                driver.find_element_by_partial_link_text("Cart").click()
                break
            else:
                action.move_to_element(item).perform()
                item.find_element_by_link_text('Add to cart').click()
                # wait for product added successfully modal
                time.sleep(5)
                driver.find_element_by_xpath("//*[@title='Continue shopping']").click()
                count += 1
                print("Product added to cart successfully")
        except Exception as e:
            print("Sorry, couldn't add " + item.text + " product to the cart")

    # get total amount on checkout screen
    total_price = driver.find_element_by_id('total_price').text

    # remove $ from total price
    # converted the total price string value to float and then round off the total price
    total_price = round(float(total_price[1:]))

    budget_amount = round(float(excel_data().cell(2, 10).value))
    # verify if user has budget more than the total price and raise assertion accordingly
    if total_price <= budget_amount:
        # click on proceed to checkout button on summary screen
        driver.find_element_by_partial_link_text("Proceed to checkout").click()

        # click on proceed to checkout button on address screen
        driver.find_element_by_name("processAddress").click()

        # select terms & conditions checkbox and click on proceed to checkout button on shipping screen
        driver.find_element_by_id('cgv').click()
        driver.find_element_by_name("processCarrier").click()

        # select mode of payment
        driver.find_element_by_class_name("cheque").click()

        # click on proceed to checkout button on payment screen
        driver.find_element_by_xpath("//button[contains(.,'I confirm my order')]").click()
        time.sleep(10)
        order_confirmation_screen_text = driver.find_element_by_class_name("navigation_page").text

        # verify if user is navigated to order confirmation screen or not and raise assertion accordingly
        if 'Order confirmation' == order_confirmation_screen_text:
            # find order confirmation text and verify accordingly
            order_confirmation_text = driver.find_element_by_xpath(
                "//h1[@class='page-heading']//following-sibling::*[contains(.,'Your order on My Store is complete.')]")
            if "Your order on My Store is complete." != order_confirmation_text.text:
                print("Couldn't process your order")
            else:
                print("Order placed successfully")
                success_screenshot_dir = os.path.join(screenshot_dir, 'Success', 'OrderConfirmation_' + fileName)
                driver.execute_script("return arguments[0].scrollIntoView();", order_confirmation_text)
                take_screenshot(success_screenshot_dir)
        else:
            print("Couldn't process your order")
    else:
        failed_screenshot_dir = os.path.join(screenshot_dir, 'Failure', 'FailedOrder_' + fileName)
        total_price_ele = driver.find_element_by_id('total_price')
        driver.execute_script("return arguments[0].scrollIntoView();", total_price_ele)
        take_screenshot(failed_screenshot_dir)
        print(
            'Order cannot be placed total amount is exceeding your budget. You need $' + str(
                total_price - budget_amount) + ' more to place your first order')

    # clear column data
    modify_excel_data(["", ""])
    
    # close the all related windows
    tearDown()


# execute test case
if __name__ == '__main__':
    unittest.main()
